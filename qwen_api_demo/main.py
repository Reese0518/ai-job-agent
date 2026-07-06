from fastapi import FastAPI
from pydantic import BaseModel
from langchain_ollama import ChatOllama
import json
from typing import List, Optional
import json
import re
from io import BytesIO
from fastapi import UploadFile, File, HTTPException, Form
from pypdf import PdfReader
from docx import Document
import sqlite3
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import FileResponse
from playwright.sync_api import sync_playwright
from urllib.parse import urljoin
from pathlib import Path


app = FastAPI(title="Qwen Local API Demo")

llm = ChatOllama(
    model="qwen2.5:7b",
    temperature=0.7,
)

class ChatRequest(BaseModel):
    message: str

class ChatResponse(BaseModel):
    answer: str

@app.get("/")
def home():
    return {"message": "Qwen API is running"}

@app.post("/chat", response_model=ChatResponse)
def chat(req: ChatRequest):
    response = llm.invoke(req.message)
    return ChatResponse(answer=response.content)

class ExtractRequest(BaseModel):
    text: str

class ResumeInfo(BaseModel):
    name: Optional[str] = None
    skills: List[str] = []
    experience_years: Optional[str] = None
    projects: List[str] = []
    strengths: List[str] = []

@app.post("/extract")
def extract(req: ExtractRequest):
    prompt = f"""
请从下面文本中提取简历信息，只返回 JSON，不要解释，不要使用 Markdown。

返回格式：
{{
  "name": "姓名，未知则为 null",
  "skills": ["技能1", "技能2"],
  "experience_years": "工作年限，未知则为 null",
  "projects": ["项目1", "项目2"],
  "strengths": ["亮点1", "亮点2"]
}}

文本：
{req.text}
"""
    response = llm.invoke(prompt)
    return json.loads(response.content)


def clean_json_text(text: str) -> str:
    text = text.strip()
    text = re.sub(r"^```json\s*", "", text)
    text = re.sub(r"^```\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    return text.strip()


def read_resume_file(file: UploadFile) -> str:
    filename = file.filename.lower()
    content = file.file.read()

    if filename.endswith(".txt"):
        return content.decode("utf-8", errors="ignore")

    if filename.endswith(".pdf"):
        reader = PdfReader(BytesIO(content))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        return "\n".join(pages)

    if filename.endswith(".docx"):
        doc = Document(BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs)

    raise HTTPException(status_code=400, detail="只支持 txt、pdf、docx 文件")


@app.post("/extract_resume_file")
async def extract_resume_file(file: UploadFile = File(...)):
    resume_text = read_resume_file(file)

    if not resume_text.strip():
        raise HTTPException(status_code=400, detail="没有读取到简历文本")

    prompt = f"""
请从下面的简历文本中提取关键信息，只返回 JSON，不要解释，不要使用 Markdown。

返回格式：
{{
  "name": "姓名，未知则为 null",
  "phone": "手机号，未知则为 null",
  "email": "邮箱，未知则为 null",
  "education": "学历信息，未知则为 null",
  "experience_years": "工作年限，未知则为 null",
  "skills": ["技能1", "技能2"],
  "projects": [
    {{
      "name": "项目名称",
      "description": "项目描述",
      "technologies": ["技术1", "技术2"]
    }}
  ],
  "work_experience": ["工作经历1", "工作经历2"],
  "strengths": ["亮点1", "亮点2"],
  "risks": ["可能的短板1", "可能的短板2"]
}}

简历文本：
{resume_text}
"""

    response = llm.invoke(prompt)
    json_text = clean_json_text(response.content)

    try:
        return json.loads(json_text)
    except json.JSONDecodeError:
        return {
            "error": "模型返回的不是合法 JSON",
            "raw_output": response.content
        }



DB_PATH = "jobs.db"


class JobUrlRequest(BaseModel):
    url: str
    resume_info: dict | None = None


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
    CREATE TABLE IF NOT EXISTS jobs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        url TEXT,
        jd_text TEXT,
        analysis_json TEXT,
        created_at TEXT
    )
    """)
    conn.commit()
    conn.close()


init_db()


def fetch_url_text(url: str) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0"
    }
    resp = requests.get(url, headers=headers, timeout=15)
    resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "html.parser")

    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()

    text = soup.get_text("\n")
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return "\n".join(lines)


@app.post("/analyze_job_url")
def analyze_job_url(req: JobUrlRequest):
    jd_text = fetch_url_text(req.url)

    if len(jd_text) < 100:
        return {
            "error": "读取到的页面文本太少，可能需要登录、验证码，或页面是动态加载的。",
            "url": req.url,
            "text_preview": jd_text[:500]
        }

    resume_part = req.resume_info or {}

    prompt = f"""
你是一个专业的招聘岗位分析 Agent。

请根据【岗位页面文本】和【候选人简历信息】进行分析，只返回 JSON，不要解释，不要使用 Markdown。

返回格式：
{{
  "job_title": "岗位名称，未知则为 null",
  "company": "公司名称，未知则为 null",
  "city": "城市，未知则为 null",
  "salary": "薪资，未知则为 null",
  "required_skills": ["必备技能1", "必备技能2"],
  "bonus_skills": ["加分项1", "加分项2"],
  "responsibilities": ["职责1", "职责2"],
  "match_score": 0,
  "match_level": "高匹配/较匹配/一般/不匹配",
  "matched_points": ["匹配点1", "匹配点2"],
  "missing_points": ["缺口1", "缺口2"],
  "resume_suggestions": ["简历优化建议1", "简历优化建议2"],
  "interview_questions": ["可能面试题1", "可能面试题2"],
  "apply_suggestion": "建议投递/谨慎投递/不建议投递"
}}

候选人简历信息：
{json.dumps(resume_part, ensure_ascii=False)}

岗位页面文本：
{jd_text[:8000]}
"""

    response = llm.invoke(prompt)
    json_text = clean_json_text(response.content)

    try:
        analysis = json.loads(json_text)
    except json.JSONDecodeError:
        analysis = {
            "error": "模型返回的不是合法 JSON",
            "raw_output": response.content
        }
    analysis = normalize_match_analysis(analysis)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.execute(
        "INSERT INTO jobs (url, jd_text, analysis_json, created_at) VALUES (?, ?, ?, ?)",
        (
            req.url,
            jd_text,
            json.dumps(analysis, ensure_ascii=False),
            datetime.now().isoformat(timespec="seconds")
        )
    )
    conn.commit()
    job_id = cursor.lastrowid
    conn.close()

    return {
        "job_id": job_id,
        "url": req.url,
        "analysis": analysis
    }


@app.get("/jobs")
def list_jobs():
    conn = sqlite3.connect(DB_PATH)
    rows = conn.execute(
        "SELECT id, url, analysis_json, created_at FROM jobs ORDER BY id DESC"
    ).fetchall()
    conn.close()

    return [
        {
            "id": row[0],
            "url": row[1],
            "analysis": json.loads(row[2]),
            "created_at": row[3]
        }
        for row in rows
    ]


@app.get("/jobs/{job_id}")
def get_job(job_id: int):
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute(
        "SELECT id, url, jd_text, analysis_json, created_at FROM jobs WHERE id = ?",
        (job_id,)
    ).fetchone()
    conn.close()

    if not row:
        return {"error": "找不到这个岗位"}

    return {
        "id": row[0],
        "url": row[1],
        "jd_text": row[2],
        "analysis": json.loads(row[3]),
        "created_at": row[4]
    }

def extract_resume_info_from_text(resume_text: str) -> dict:
    prompt = f"""
请从下面的简历文本中提取关键信息，只返回 JSON，不要解释，不要使用 Markdown。

返回格式：
{{
  "name": "姓名，未知则为 null",
  "phone": "手机号，未知则为 null",
  "email": "邮箱，未知则为 null",
  "education": "学历信息，未知则为 null",
  "experience_years": "工作年限，未知则为 null",
  "skills": ["技能1", "技能2"],
  "projects": [
    {{
      "name": "项目名称",
      "description": "项目描述",
      "technologies": ["技术1", "技术2"]
    }}
  ],
  "work_experience": ["工作经历1", "工作经历2"],
  "strengths": ["亮点1", "亮点2"],
  "risks": ["可能的短板1", "可能的短板2"]
}}

简历文本：
{resume_text}
"""
    response = llm.invoke(prompt)
    json_text = clean_json_text(response.content)

    try:
        return json.loads(json_text)
    except json.JSONDecodeError:
        return {
            "error": "简历提取结果不是合法 JSON",
            "raw_output": response.content
        }


@app.post("/analyze_job_with_resume_file")
async def analyze_job_with_resume_file(
    job_url: str = Form(...),
    file: UploadFile = File(...)
):
    resume_text = read_resume_file(file)

    if not resume_text.strip():
        raise HTTPException(status_code=400, detail="没有读取到简历文本")

    resume_info = extract_resume_info_from_text(resume_text)

    if "error" in resume_info:
        return {
            "error": "简历信息提取失败",
            "resume_info": resume_info
        }

    jd_text = fetch_url_text(job_url)

    if len(jd_text) < 100:
        return {
            "error": "读取到的 JD 文本太少，可能需要登录、验证码，或页面是动态加载的。",
            "url": job_url,
            "text_preview": jd_text[:500]
        }

    prompt = f"""
你是一个专业的岗位匹配分析 Agent。请基于证据分析候选人与岗位的匹配度。

重要规则：
1. 必须同时阅读【候选人简历信息】和【岗位JD文本】。
2. 不要只看关键词完全相同，要识别同义技能：
   - RAG = 知识库问答 / 检索增强生成 / 向量检索
   - FastAPI = API接口开发 / 后端服务
   - Qwen/Ollama = 本地大模型调用 / 模型部署实践
   - LangChain = LLM应用编排 / Agent开发基础
   - Streamlit = AI应用原型界面
3. 如果候选人有相关项目经验，应给出“部分匹配”或“较匹配”，不要直接判定没有技能。
4. 如果岗位要求商业经验，而候选人是学习/项目经验，要标为“经验不足”，不是“完全不匹配”。
5. 每个匹配点和缺口都必须写出证据来源。
6. 只返回 JSON，不要解释，不要使用 Markdown。

返回格式：
{{
  "job_title": "岗位名称，未知则为 null",
  "company": "公司名称，未知则为 null",
  "city": "城市，未知则为 null",
  "salary": "薪资，未知则为 null",
  "match_score": 0,
  "match_level": "高匹配/较匹配/一般/不匹配",
  "required_skills": ["必备技能1", "必备技能2"],
  "bonus_skills": ["加分项1", "加分项2"],
  "matched_points": [
    {{
      "point": "匹配点",
      "resume_evidence": "简历证据",
      "jd_evidence": "JD证据"
    }}
  ],
  "missing_points": [
    {{
      "point": "缺口",
      "reason": "为什么是缺口",
      "improvement": "如何补强"
    }}
  ],
  "skill_mapping": [
    {{
      "jd_skill": "JD中的技能",
      "resume_skill": "简历中的对应技能",
      "match_type": "完全匹配/同义匹配/部分匹配/缺失"
    }}
  ],
  "resume_suggestions": ["简历优化建议1", "简历优化建议2"],
  "interview_questions": ["可能面试题1", "可能面试题2"],
  "apply_suggestion": "建议投递/谨慎投递/不建议投递"
}}

候选人简历信息：
{json.dumps(resume_info, ensure_ascii=False)}

岗位JD文本：
{jd_text[:8000]}
"""

    response = llm.invoke(prompt)
    json_text = clean_json_text(response.content)

    try:
        analysis = json.loads(json_text)
    except json.JSONDecodeError:
        analysis = {
            "error": "岗位分析结果不是合法 JSON",
            "raw_output": response.content
        }
    analysis = normalize_match_analysis(analysis)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.execute(
        "INSERT INTO jobs (url, jd_text, analysis_json, created_at) VALUES (?, ?, ?, ?)",
        (
            job_url,
            jd_text,
            json.dumps(
                {
                    "resume_info": resume_info,
                    "analysis": analysis
                },
                ensure_ascii=False
            ),
            datetime.now().isoformat(timespec="seconds")
        )
    )
    conn.commit()
    job_id = cursor.lastrowid
    conn.close()

    return {
        "job_id": job_id,
        "job_url": job_url,
        "resume_info": resume_info,
        "analysis": analysis
    }
@app.get("/export_jobs_excel")
def export_jobs_excel():
    conn = sqlite3.connect(DB_PATH)
    rows = conn.execute(
        "SELECT id, url, analysis_json, created_at FROM jobs ORDER BY id DESC"
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "岗位分析结果"

    headers = [
        "ID",
        "岗位链接",
        "岗位名称",
        "公司",
        "城市",
        "薪资",
        "匹配分",
        "匹配等级",
        "投递建议",
        "匹配点",
        "缺口",
        "简历优化建议",
        "面试问题",
        "创建时间",
    ]
    ws.append(headers)

    for row in rows:
        job_id, url, analysis_json, created_at = row
        data = json.loads(analysis_json)

        analysis = data.get("analysis", data)

        matched_points = analysis.get("matched_points", [])
        missing_points = analysis.get("missing_points", [])
        suggestions = analysis.get("resume_suggestions", [])
        questions = analysis.get("interview_questions", [])

        def to_text(value):
            if isinstance(value, list):
                return "\n".join(
                    json.dumps(item, ensure_ascii=False) if isinstance(item, dict) else str(item)
                    for item in value
                )
            if isinstance(value, dict):
                return json.dumps(value, ensure_ascii=False)
            return value or ""

        ws.append([
            job_id,
            url,
            analysis.get("job_title", ""),
            analysis.get("company", ""),
            analysis.get("city", ""),
            analysis.get("salary", ""),
            analysis.get("match_score", ""),
            analysis.get("match_level", ""),
            analysis.get("apply_suggestion", ""),
            to_text(matched_points),
            to_text(missing_points),
            to_text(suggestions),
            to_text(questions),
            created_at,
        ])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_widths = {
        "A": 8,
        "B": 45,
        "C": 24,
        "D": 24,
        "E": 12,
        "F": 14,
        "G": 10,
        "H": 14,
        "I": 16,
        "J": 45,
        "K": 45,
        "L": 45,
        "M": 45,
        "N": 20,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    output_path = "job_analysis_results.xlsx"
    wb.save(output_path)

    return FileResponse(
        output_path,
        filename="job_analysis_results.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


class DynamicCrawlRequest(BaseModel):
    url: str
    wait_seconds: int = 5


@app.post("/crawl_dynamic_job_links")
def crawl_dynamic_job_links(req: DynamicCrawlRequest):
    job_keywords = [
        "job", "jobs", "career", "position", "recruit",
        "职位", "岗位", "招聘", "详情"
    ]

    results = []
    seen = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(
            user_agent="Mozilla/5.0"
        )

        page.goto(req.url, wait_until="networkidle", timeout=30000)
        page.wait_for_timeout(req.wait_seconds * 1000)

        links = page.locator("a").evaluate_all("""
        els => els.map(a => ({
            text: a.innerText || "",
            href: a.href || ""
        }))
        """)

        browser.close()

    for item in links:
        text = item.get("text", "").strip()
        href = item.get("href", "").strip()

        if not href:
            continue

        combined = (text + " " + href).lower()

        if any(k.lower() in combined for k in job_keywords):
            full_url = urljoin(req.url, href)

            if full_url not in seen:
                seen.add(full_url)
                results.append({
                    "text": text,
                    "url": full_url
                })

    return {
        "source_url": req.url,
        "count": len(results),
        "links": results
    }

@app.post("/debug_dynamic_page")
def debug_dynamic_page(req: DynamicCrawlRequest):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(user_agent="Mozilla/5.0")

        page.goto(req.url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(req.wait_seconds * 1000)

        title = page.title()

        try:
            text = page.locator("body").inner_text(timeout=5000)
        except Exception as e:
            text = f"读取 body 文本失败: {e}"

        links = page.locator("a").evaluate_all("""
        els => els.slice(0, 50).map(a => ({
            text: a.innerText || "",
            href: a.href || ""
        }))
        """)

        html = page.content()
        browser.close()

    return {
        "url": req.url,
        "title": title,
        "text_preview": text[:2000],
        "html_length": len(html),
        "link_count_preview": len(links),
        "links_preview": links
    }


def analyze_jd_text_with_resume_info(jd_text: str, resume_info: dict) -> dict:
    prompt = f"""
你是一个专业的岗位匹配分析 Agent。
请根据候选人简历信息和岗位 JD，判断候选人与岗位是否匹配。

重要规则：
1. 不要只做关键词完全匹配，要识别同义技能和相关经验。
2. 如果简历里有项目经验、实习经验、课程项目，也可以作为匹配依据。
3. 如果 JD 要求的技能在简历中没有直接出现，但有相近技能，请写入 skill_mapping。
4. 只返回 JSON，不要解释，不要使用 Markdown。

返回格式：
{{
  "job_title": "岗位名称",
  "company": "公司名称",
  "city": "城市",
  "salary": "薪资",
  "match_score": 0,
  "match_level": "高/中/低",
  "apply_suggestion": "建议投递/谨慎投递/不建议投递",
  "matched_points": [
    {{"point": "匹配点", "evidence": "简历或JD中的依据"}}
  ],
  "missing_points": [
    {{"point": "缺口", "reason": "为什么是缺口", "improvement": "如何补强"}}
  ],
  "skill_mapping": [
    {{"jd_skill": "JD技能", "resume_skill": "简历对应技能", "match_type": "完全匹配/同义匹配/部分匹配/缺失"}}
  ],
  "resume_suggestions": ["简历优化建议"],
  "interview_questions": ["可能面试问题"]
}}

候选人简历信息：
{json.dumps(resume_info, ensure_ascii=False)}

岗位 JD：
{jd_text[:8000]}
"""

    response = llm.invoke(prompt)
    json_text = clean_json_text(response.content)

    try:
        return normalize_match_analysis(json.loads(json_text))
    except json.JSONDecodeError:
        return normalize_match_analysis({
            "error": "岗位分析结果不是合法 JSON",
            "raw_output": response.content
        })


def save_job_analysis(url: str, jd_text: str, resume_info: dict, analysis: dict) -> int:
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.execute(
        "INSERT INTO jobs (url, jd_text, analysis_json, created_at) VALUES (?, ?, ?, ?)",
        (
            url,
            jd_text,
            json.dumps(
                {
                    "resume_info": resume_info,
                    "analysis": analysis
                },
                ensure_ascii=False
            ),
            datetime.now().isoformat(timespec="seconds")
        )
    )
    conn.commit()
    job_id = cursor.lastrowid
    conn.close()
    return job_id


def normalize_match_analysis(analysis: dict) -> dict:
    if not isinstance(analysis, dict):
        return analysis

    raw_score = analysis.get("match_score", 0)

    try:
        if isinstance(raw_score, str):
            raw_score = raw_score.strip().replace("%", "")
        score = float(raw_score)
    except (TypeError, ValueError):
        score = 0

    if 0 < score <= 10:
        score = score * 10

    score = max(0, min(100, round(score)))
    analysis["match_score"] = score

    if score >= 80:
        analysis["match_level"] = "高"
        analysis["apply_suggestion"] = "建议投递"
    elif score >= 60:
        analysis["match_level"] = "中"
        analysis["apply_suggestion"] = "谨慎投递"
    else:
        analysis["match_level"] = "低"
        analysis["apply_suggestion"] = "不建议投递"

    return analysis


def normalize_job_links(raw_links):
    if isinstance(raw_links, dict):
        raw_links = raw_links.get("links", [])

    normalized = []
    for item in raw_links:
        if isinstance(item, str):
            normalized.append({"title": "", "url": item})
        elif isinstance(item, dict):
            normalized.append({
                "title": item.get("title") or item.get("text") or "",
                "url": item.get("url") or item.get("href") or ""
            })
    return normalized


def read_job_detail_with_browser(page, url: str) -> str:
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(2500)

    try:
        page.wait_for_load_state("networkidle", timeout=8000)
    except Exception:
        pass

    selectors = [
        ".job-detail",
        ".job-detail-box",
        ".job-info",
        ".position-detail",
        ".detail-content",
        "main",
        "body"
    ]

    for selector in selectors:
        try:
            text = page.locator(selector).first.inner_text(timeout=3000).strip()
            if len(text) >= 100:
                return text
        except Exception:
            continue

    try:
        return page.locator("body").inner_text(timeout=5000).strip()
    except Exception:
        return ""


def build_browser_context(playwright):
    state_path = Path("zhaopin_state.json")
    browser = playwright.chromium.launch(headless=True)

    context_kwargs = {
        "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36",
        "viewport": {"width": 1366, "height": 900}
    }

    if state_path.exists():
        context_kwargs["storage_state"] = str(state_path)

    context = browser.new_context(**context_kwargs)
    return browser, context


@app.post("/batch_analyze_job_links_with_resume_file")
def batch_analyze_job_links_with_resume_file(
    resume_file: UploadFile = File(...),
    links_file: UploadFile = File(...),
    start_index: int = Form(1),
    max_jobs: int = Form(5),
    analyze: bool = Form(True)
):
    resume_text = read_resume_file(resume_file)
    if not resume_text.strip():
        raise HTTPException(status_code=400, detail="没有读取到简历文本")

    resume_info = extract_resume_info_from_text(resume_text)

    links_content = links_file.file.read()
    try:
        raw_links = json.loads(links_content.decode("utf-8-sig"))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"链接文件不是合法 JSON: {e}")

    job_links = normalize_job_links(raw_links)
    if not job_links:
        raise HTTPException(status_code=400, detail="链接文件里没有可分析的岗位链接")

    if start_index < 1:
        start_index = 1
    if max_jobs < 1:
        max_jobs = 5
    if max_jobs > 20:
        max_jobs = 20

    total_links = len(job_links)
    selected_links = job_links[start_index - 1:start_index - 1 + max_jobs]

    results = []
    failed = []

    with sync_playwright() as p:
        browser, context = build_browser_context(p)
        page = context.new_page()

        try:
            for offset, item in enumerate(selected_links, start=0):
                index = start_index + offset
                title = item.get("title", "")
                url = item.get("url", "")

                if not url:
                    failed.append({
                        "index": index,
                        "title": title,
                        "url": url,
                        "reason": "缺少 url"
                    })
                    continue

                try:
                    read_source = "requests"
                    request_error = ""

                    try:
                        jd_text = fetch_url_text(url)
                    except Exception as e:
                        jd_text = ""
                        request_error = str(e)

                    if len(jd_text) < 100:
                        read_source = "browser"
                        jd_text = read_job_detail_with_browser(page, url)

                    if len(jd_text) < 100:
                        failed.append({
                            "index": index,
                            "title": title,
                            "url": url,
                            "reason": "没有读取到足够的 JD 文本，普通请求和浏览器读取都失败",
                            "request_error": request_error,
                            "browser_text_preview": jd_text[:200]
                        })
                        continue

                    if analyze:
                        analysis = analyze_jd_text_with_resume_info(jd_text, resume_info)
                    else:
                        analysis = {
                            "job_title": title,
                            "match_score": "",
                            "match_level": "",
                            "apply_suggestion": "仅抓取JD，未调用模型分析"
                        }

                    if title and not analysis.get("job_title"):
                        analysis["job_title"] = title

                    job_id = save_job_analysis(url, jd_text, resume_info, analysis)

                    results.append({
                        "index": index,
                        "job_id": job_id,
                        "title": title or analysis.get("job_title", ""),
                        "url": url,
                        "read_source": read_source,
                        "match_score": analysis.get("match_score", ""),
                        "match_level": analysis.get("match_level", ""),
                        "apply_suggestion": analysis.get("apply_suggestion", "")
                    })
                except Exception as e:
                    failed.append({
                        "index": index,
                        "title": title,
                        "url": url,
                        "reason": str(e)
                    })
        finally:
            context.close()
            browser.close()

    return {
        "total": total_links,
        "start_index": start_index,
        "max_jobs": max_jobs,
        "processed": len(selected_links),
        "analyze": analyze,
        "success": len(results),
        "failed_count": len(failed),
        "results": results,
        "failed": failed,
        "next_step": "分析完成后，可以调用 /export_jobs_excel 下载 Excel"
    }
